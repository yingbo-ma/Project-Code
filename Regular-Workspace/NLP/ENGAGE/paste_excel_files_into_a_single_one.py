# importing openpyxl module
import openpyxl as xl
import os

data_folder_path = "C:\\Users\\Yingbo\\Desktop\\ENGAGE dataset\\ENGAGE Coded Transcriptions\\Previous\\paste_with_values\\"
file_name_arr = os.listdir(data_folder_path)
print(file_name_arr)
print(len(file_name_arr))

target_file = "C:\\Users\\Yingbo\\Desktop\\Corpus.xlsx"

for file_index in range(len(file_name_arr)):

    # opening the source excel file
    origin_filename = file_name_arr[file_index]
    print("reading from ", origin_filename, "...")
    wb1 = xl.load_workbook(data_folder_path+origin_filename)
    ws1 = wb1.worksheets[0]

    # opening the destination excel file
    wb2 = xl.load_workbook(target_file)
    ws2 = wb2.active

    # calculate total number of rows and columns in source excel file
    mr1 = ws1.max_row
    mc1 = ws1.max_column

    # calculate total number of rows and columns in target excel file
    mr2 = ws2.max_row
    mc2 = ws2.max_column

    print("copying starting from row ", mr2)

    # copying the cell values from source
    # excel file to destination excel file
    for i in range(2, mr1 + 1):
        for j in range(1, mc1 + 1):
            # reading cell value from source excel file
            c = ws1.cell(row=i, column=j)

            # writing the read value to destination excel file
            ws2.cell(row=mr2+i-1, column=j).value = c.value

    # saving the destination excel file
    wb2.save(str(target_file))